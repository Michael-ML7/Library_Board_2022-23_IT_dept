# import student_name_to_id
from datetime import time
from openpyxl import Workbook, load_workbook
# import datetime

i = 5

wb = load_workbook('2022-2023 Library Board Attendance (Full day school, lunch & after school).xlsx')
ws = wb["Automatic System Record"]

def col_int_to_word(n):
    convertString = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    base = 26
    i = n - 1
    if i < base:
        return convertString[i]
    else:
        return col_int_to_word(i // base) + convertString[i % base]

while ws["C%s" %(i + 1)].value != None:
    # print(ws["C%s" %(i + 1)].value)
    i += 1

    student_id = str(ws["D%s" %i].value)
    student_id = "s" + student_id + "@lsc.hk"
    time_arrived = ws["E%s" %i].value
    curr_date = ws["F%s" %i].value
    # print(student_id)
    # print(time_arrived)
    # print(curr_date)
    # print(curr_date.weekday())

    day_of_week = curr_date.weekday()
    day_of_week_str = ""
    if day_of_week == 0:
        day_of_week_str = "Monday"
    elif day_of_week == 1:
        day_of_week_str = "Tuesday"
    elif day_of_week == 2:
        day_of_week_str = "Wednesday"
    elif day_of_week == 3:
        day_of_week_str = "Thursday"
    elif day_of_week == 4:
        day_of_week_str = "Friday"

    # print(day_of_week_str)

    duty_slot = "Lunch"
    if time_arrived >= time(14, 55, 00):
        duty_slot = "Afterschool"

    # print("%s %s" %(day_of_week_str, duty_slot))
    attendance_sheet = wb["%s %s" %(day_of_week_str, duty_slot)]

    row = 1
    while attendance_sheet["C%s" %row].value != None:
        if attendance_sheet["C%s" %row].value == student_id:
            break
        row += 1

    bonus = False
    if attendance_sheet["C%s" %row].value != student_id:
        bonus = True
    
    if not bonus:
        col = 7
        while attendance_sheet["%s1" %(col_int_to_word(col))].value != None:
            if attendance_sheet["%s1" %(col_int_to_word(col))].value == curr_date:
                break
            col += 1
        
        if attendance_sheet["%s1" %(col_int_to_word(col))] != curr_date:
            print("ERROR")
        
        attendance_sheet["%s%s" %(col_int_to_word(col), row)].value = "TRUE"
    else:
        attendance_sheet = wb["Bonus"]
        row = 3
        while attendance_sheet["C%s" %row] != student_id:
            row += 1
        col = 7
        while attendance_sheet["%s%s" %(col_int_to_word(col), row)] != None:
            col += 1
        attendance_sheet["%s%s" %(col_int_to_word(col), row)] = curr_date