# import student_name_to_id
from datetime import time
from openpyxl import Workbook, load_workbook
# import datetime

wb = load_workbook('22-23 Library Board Student ID Collection.xlsx')
ws = wb["Sheet1"]

def col_int_to_word(n):
    convertString = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    base = 26
    i = n - 1
    if i < base:
        return convertString[i]
    else:
        return col_int_to_word(i // base) + convertString[i % base]

file = open("student_card_code.txt", "w")

barcode = []
student_id = []

for i in range(2, 80):
    barcode.append(ws["F%s" %i].value)
    student_id.append(ws["G%s" %i].value)

for i in range(0, 78):
    file.write(barcode[i])
    file.write(' ')
    file.write(student_id[i])
    file.write('\n')