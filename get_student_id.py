from openpyxl import Workbook, load_workbook

wb = load_workbook('2022-2023 Library Board Attendance (Full day school, lunch & after school).xlsx')
ws = wb["Master Attendance "]

a = []

for i in range(9, 87):
    a.append(ws["C%s" %i].value)

f = open("student_id.txt", "w")
for x in a:
    f.write(x)
    f.write('\n')