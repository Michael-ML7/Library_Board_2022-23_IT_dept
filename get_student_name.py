from openpyxl import Workbook, load_workbook

wb = load_workbook('2022-2023 Library Board Attendance (Full day school, lunch & after school).xlsx')
ws = wb["Master Attendance "]

a = []

for i in range(9, 87):
    a.append(ws["D%s" %i].value)

f = open("student_name.txt", "w")
for x in a:
    s = ""
    for val in x:
        if val == ' ' or ('a' <= val and val <= 'z') or ('A' <= val and val <= 'Z'):
            s = s + val
    s = s.upper()
    f.write(s)
    f.write('\n')